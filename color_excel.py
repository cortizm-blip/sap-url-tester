#!/usr/bin/env python3
import sys, json
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

excel_path = sys.argv[1]   # input xlsx
out_path   = sys.argv[2]   # output xlsx
results_json = sys.argv[3] # JSON string with results

results = json.loads(results_json)

# Build maps: url -> result, name -> result
by_url  = {r['url'].strip(): r  for r in results if r.get('url')}
by_name = {r['name'].strip(): r for r in results if r.get('name')}

wb = openpyxl.load_workbook(excel_path)
ws = wb.active

# Detect header row columns
url_col = nombre_col = estatus_col = None
for cell in ws[1]:
    if cell.value is None: continue
    v = str(cell.value).lower().strip()
    if v == 'url':     url_col      = cell.column
    if v == 'nombre':  nombre_col   = cell.column
    if 'estatus' in v or 'status' in v: estatus_col = cell.column

if estatus_col is None: estatus_col = 4  # default col D

fill_ok   = PatternFill("solid", fgColor="C6EFCE")
fill_err  = PatternFill("solid", fgColor="FFC7CE")
font_ok   = Font(color="276221", bold=True)
font_err  = Font(color="9C0006", bold=True)
align_c   = Alignment(horizontal="center", vertical="center")

updated = 0
for row in ws.iter_rows(min_row=2):
    url_val    = str(row[url_col-1].value or '').strip()    if url_col    else ''
    nombre_val = str(row[nombre_col-1].value or '').strip() if nombre_col else ''

    result = by_url.get(url_val) or by_name.get(nombre_val)
    if not result:
        continue

    ok = result.get('ok', False)
    if ok:
        status_text = f"OK - HTTP {result.get('status','')}"
    elif not result.get('reachable', True):
        status_text = result.get('error_type', 'ERROR - Sin conexión')
    else:
        status_text = f"ERROR - HTTP {result.get('status','N/A')}"

    cell = row[estatus_col - 1]
    cell.value     = status_text
    cell.fill      = fill_ok  if ok else fill_err
    cell.font      = font_ok  if ok else font_err
    cell.alignment = align_c
    updated += 1

# Ajustar ancho columna estatus
ws.column_dimensions[openpyxl.utils.get_column_letter(estatus_col)].width = 25

wb.save(out_path)
print(json.dumps({"ok": True, "updated": updated}))
